#!/usr/bin/env python3
"""
Trading Journal Generator V2.2
FIXED: Parser universale per V37 + Perpetual
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import os

class TradingJournal:
    def __init__(self):
        self.v37_trades_file = 'paper_trading_30d/trades.json'
        self.perpetual_trades_file = 'perpetual_bot/perpetual_data/trades.json'
        self.output_file = 'Trading_Journal_LIVE.xlsx'
    
    def normalize_trade(self, trade, bot_type="V37"):
        """
        Normalizza formato trade (V37 usa chiavi diverse da Perpetual)
        """
        # Determina chiavi basate su bot type o auto-detect
        if 'entry_price' in trade:
            # Formato Perpetual
            return {
                'symbol': trade.get('symbol', 'N/A'),
                'direction': trade.get('direction', 'N/A'),
                'entry_price': trade.get('entry_price', 0),
                'exit_price': trade.get('exit_price', 0),
                'quantity': trade.get('quantity', 0),
                'pnl_usd': trade.get('pnl_usd', 0),
                'pnl_pct': trade.get('pnl_pct', 0),  # Decimale
                'entry_time': trade.get('entry_time', ''),
                'exit_time': trade.get('exit_time', ''),
                'exit_reason': trade.get('exit_reason', 'N/A'),
                'bot': 'Perpetual'
            }
        else:
            # Formato V37
            return {
                'symbol': trade.get('symbol', 'N/A'),
                'direction': 'N/A',  # V37 non ha direction (solo LONG)
                'entry_price': trade.get('entry', 0),
                'exit_price': trade.get('exit', 0),
                'quantity': trade.get('size', 0),
                'pnl_usd': trade.get('pnl', 0),
                'pnl_pct': trade.get('pnl_pct', 0) / 100,  # Da % a decimale
                'entry_time': '',  # V37 non ha entry_time
                'exit_time': trade.get('closed_at', ''),
                'exit_reason': trade.get('exit_reason', 'N/A'),
                'bot': 'V37'
            }
    
    def load_trades(self, filepath, bot_type="V37"):
        """Carica e normalizza trade da JSON"""
        try:
            with open(filepath) as f:
                raw_trades = json.load(f)
                # Normalizza tutti i trade
                normalized = [self.normalize_trade(t, bot_type) for t in raw_trades]
                return normalized
        except Exception as e:
            print(f"⚠️ Error loading {filepath}: {e}")
            return []
    
    def create_excel_journal(self):
        """Crea Excel journal professionale"""
        
        print("📊 Loading trades...")
        v37_trades = self.load_trades(self.v37_trades_file, "V37")
        perpetual_trades = self.load_trades(self.perpetual_trades_file, "Perpetual")
        
        print(f"   V37: {len(v37_trades)} trades")
        print(f"   Perpetual: {len(perpetual_trades)} trades")
        
        # Crea workbook
        wb = Workbook()
        
        print("📋 Creating sheets...")
        self._create_overview_sheet(wb, v37_trades, perpetual_trades)
        self._create_trades_sheet(wb, v37_trades, "V37 Spot", "V37")
        self._create_trades_sheet(wb, perpetual_trades, "Perpetual", "PERP")
        self._create_analytics_sheet(wb, v37_trades, perpetual_trades)
        
        print(f"💾 Saving to {self.output_file}...")
        wb.save(self.output_file)
        
        print("")
        print("✅ Journal updated successfully!")
        print(f"📁 File: {self.output_file}")
        
        return self.output_file
    
    def _create_overview_sheet(self, wb, v37_trades, perpetual_trades):
        """Sheet 1: Dashboard Overview"""
        ws = wb.active
        ws.title = "📊 Dashboard"
        
        # Header
        ws['A1'] = "🎯 TRADING JOURNAL - LIVE DASHBOARD"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Last Update: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=9)
        
        # V37 Stats
        row = 4
        ws[f'A{row}'] = "V37 SPOT BOT"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="4472C4", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        stats = self._calculate_stats(v37_trades)
        self._add_stat_rows(ws, row, stats)
        
        # Perpetual Stats
        row += len(stats) + 2
        ws[f'A{row}'] = "PERPETUAL BOT"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="ED7D31", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        perp_stats = self._calculate_stats(perpetual_trades)
        self._add_stat_rows(ws, row, perp_stats)
        
        # Combined
        row += len(perp_stats) + 2
        ws[f'A{row}'] = "COMBINED PORTFOLIO"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="70AD47", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        combined_stats = self._calculate_stats(v37_trades + perpetual_trades)
        self._add_stat_rows(ws, row, combined_stats)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_trades_sheet(self, wb, trades, name, prefix):
        """Sheet trade dettagliati"""
        ws = wb.create_sheet(title=name)
        
        headers = ['Trade ID', 'Symbol', 'Direction', 'Entry', 'Exit', 
                   'Quantity', 'PnL $', 'PnL %', 'Reason', 'Exit Time']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Ordina per exit_time
        sorted_trades = sorted(trades, key=lambda x: x.get('exit_time', ''), reverse=True)
        
        for idx, trade in enumerate(sorted_trades, 2):
            trade_id = f"{prefix}-{trade.get('exit_time', '')[:10].replace('-', '')}-{idx-1:03d}"
            
            ws.cell(idx, 1, trade_id)
            ws.cell(idx, 2, trade['symbol'])
            ws.cell(idx, 3, trade['direction'])
            ws.cell(idx, 4, round(trade['entry_price'], 2))
            ws.cell(idx, 5, round(trade['exit_price'], 2))
            ws.cell(idx, 6, round(trade['quantity'], 6))
            
            pnl = trade['pnl_usd']
            ws.cell(idx, 7, round(pnl, 2))
            
            # pnl_pct già normalizzato in decimale
            pnl_pct = trade['pnl_pct'] * 100
            ws.cell(idx, 8, round(pnl_pct, 2))
            
            ws.cell(idx, 9, trade['exit_reason'])
            ws.cell(idx, 10, trade['exit_time'][:19])
            
            # Colora PnL
            pnl_cell = ws.cell(idx, 7)
            pnl_pct_cell = ws.cell(idx, 8)
            
            color = "C6EFCE" if pnl > 0 else "FFC7CE" if pnl < 0 else "FFFFFF"
            pnl_cell.fill = PatternFill(start_color=color, fill_type="solid")
            pnl_pct_cell.fill = PatternFill(start_color=color, fill_type="solid")
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    def _create_analytics_sheet(self, wb, v37_trades, perpetual_trades):
        """Sheet analytics"""
        ws = wb.create_sheet(title="📈 Analytics")
        
        ws['A1'] = "EQUITY CURVE - COMBINED PORTFOLIO"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        all_trades = v37_trades + perpetual_trades
        all_trades.sort(key=lambda x: x.get('exit_time', ''))
        
        if not all_trades:
            ws['A3'] = "No trades yet"
            return
        
        row = 3
        headers = ['Trade #', 'Symbol', 'Bot', 'PnL $', 'Cumulative $']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
        
        cumulative = 0
        for idx, trade in enumerate(all_trades, 1):
            pnl = trade['pnl_usd']
            cumulative += pnl
            
            ws.cell(row + idx, 1, idx)
            ws.cell(row + idx, 2, trade['symbol'])
            ws.cell(row + idx, 3, trade['bot'])
            ws.cell(row + idx, 4, round(pnl, 2))
            ws.cell(row + idx, 5, round(cumulative, 2))
            
            cum_cell = ws.cell(row + idx, 5)
            if cumulative > 0:
                cum_cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
            elif cumulative < 0:
                cum_cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        
        if len(all_trades) > 1:
            chart = LineChart()
            chart.title = "Cumulative PnL Over Time"
            chart.y_axis.title = "Cumulative PnL ($)"
            chart.x_axis.title = "Trade #"
            
            data = Reference(ws, min_col=5, min_row=row, max_row=row + len(all_trades))
            cats = Reference(ws, min_col=1, min_row=row+1, max_row=row + len(all_trades))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            chart.height = 10
            chart.width = 20
            
            ws.add_chart(chart, "G3")
    
    def _calculate_stats(self, trades):
        """Calcola statistiche"""
        if not trades:
            return {
                'Total Trades': 0,
                'Winning Trades': 0,
                'Losing Trades': 0,
                'Win Rate': '0.0%',
                'Total PnL': '$0.00',
                'Avg PnL': '$0.00',
                'Best Trade': '$0.00',
                'Worst Trade': '$0.00',
                'Sharpe Estimate': '0.00'
            }
        
        total = len(trades)
        wins = sum(1 for t in trades if t['pnl_usd'] > 0)
        losses = total - wins
        win_rate = (wins / total * 100) if total > 0 else 0
        
        pnls = [t['pnl_usd'] for t in trades]
        total_pnl = sum(pnls)
        avg_pnl = total_pnl / total if total > 0 else 0
        
        import math
        if len(pnls) > 1:
            std_dev = math.sqrt(sum((x - avg_pnl) ** 2 for x in pnls) / len(pnls))
            sharpe = (avg_pnl / std_dev * math.sqrt(252)) if std_dev > 0 else 0
        else:
            sharpe = 0
        
        return {
            'Total Trades': total,
            'Winning Trades': wins,
            'Losing Trades': losses,
            'Win Rate': f'{win_rate:.1f}%',
            'Total PnL': f'${total_pnl:.2f}',
            'Avg PnL': f'${avg_pnl:.2f}',
            'Best Trade': f'${max(pnls):.2f}',
            'Worst Trade': f'${min(pnls):.2f}',
            'Sharpe Estimate': f'{sharpe:.2f}'
        }
    
    def _add_stat_rows(self, ws, start_row, stats):
        """Aggiungi righe statistiche"""
        for idx, (key, value) in enumerate(stats.items()):
            row = start_row + idx
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            
            if 'Win Rate' in key:
                ws[f'B{row}'].fill = PatternFill(start_color="FFEB9C", fill_type="solid")
            elif 'Total PnL' in key:
                if '$-' in str(value):
                    ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                else:
                    ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", fill_type="solid")

if __name__ == "__main__":
    print("════════════════════════════════════════════════════════════")
    print("📊 TRADING JOURNAL V2.2 - UNIVERSAL PARSER")
    print("════════════════════════════════════════════════════════════")
    print("")
    
    journal = TradingJournal()
    journal.create_excel_journal()
    
    print("")
    print("════════════════════════════════════════════════════════════")
